import json
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# TileProcessor Class: extract_tileset.py functionality
class TileProcessor:
    def __init__(self, file_path, output_path):
        self.file_path = file_path
        self.output_path = output_path
        self.data = None
        self.tiles_df_sorted = None
        self.tiles_df_simplified = None
        self.tiles_df_reordered = None

    def load_json(self):
        with open(self.file_path, 'r') as f:
            self.data = json.load(f)

    def fill_missing_tiles(self):
        width, height, length = self.data.get("width", 0), self.data.get("height", 0), self.data.get("length", 0)
        tiles = self.data.get("tiles", {})
        for x in range(width):
            for y in range(height):
                for z in range(length):
                    key = f"{x},{y},{z}"
                    if key not in tiles and z == 0:
                        tiles[key] = {
                            "x": x, "y": y, "z": z,
                            "tileType": {"image": "tile-empty.png"},
                            "items": {"rampPoints": False},
                            "underRamp": False
                        }
        self.data["tiles"] = tiles

    def extract_and_sort_tiles(self):
        tiles = self.data.get("tiles", {})
        tiles_df = pd.DataFrame.from_dict(tiles, orient='index')
        self.tiles_df_sorted = tiles_df[["x", "y", "z", "tileType", "items"]].sort_values(by=["x", "y", "z"])

    def process_tiles(self):
        self.tiles_df_simplified = self.tiles_df_sorted[["x", "y", "z"]].copy()
        self.tiles_df_simplified["tileType_image"] = self.tiles_df_sorted["tileType"].apply(
            lambda x: x.get("image") if isinstance(x, dict) else None)
        self.tiles_df_simplified["rampPoints"] = self.tiles_df_sorted["items"].apply(
            lambda x: x.get("rampPoints") if isinstance(x, dict) else None)
        self.tiles_df_simplified["underRamp"] = self.tiles_df_simplified.apply(
            lambda row: row["z"] == 0 and self.tiles_df_simplified.query("x==@row.x & y==@row.y").shape[0] > 1, axis=1)

    def reorder_tiles(self):
        ev_rows = self.tiles_df_simplified[self.tiles_df_simplified["tileType_image"].isin(["ev1.png", "ev2.png", "ev3.png"])]
        non_ev_rows = self.tiles_df_simplified[~self.tiles_df_simplified["tileType_image"].isin(["ev1.png", "ev2.png", "ev3.png"])]
        underramp_rows = non_ev_rows[non_ev_rows["underRamp"]]
        ramppoint_rows = non_ev_rows[non_ev_rows["rampPoints"]]
        other_rows = non_ev_rows[~non_ev_rows.index.isin(underramp_rows.index.union(ramppoint_rows.index))]
        self.tiles_df_reordered = pd.concat([other_rows, ramppoint_rows, underramp_rows, ev_rows])

    def save_to_excel(self, xlsx_output):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tileset Data"

        ws.append(["x", "y", "z", "tileType_image", "rampPoints", "underRamp"])
        for _, row in self.tiles_df_reordered.iterrows():
            ws.append(row.tolist())

        wb.save(xlsx_output)
        print(f"Data saved to {xlsx_output}")

    def process(self, xlsx_output):
        self.load_json()
        self.fill_missing_tiles()
        self.extract_and_sort_tiles()
        self.process_tiles()
        self.reorder_tiles()
        self.save_to_excel(xlsx_output)


# TileImageProcessor Class: count_tileset.py functionality
class TileImageProcessor:
    def __init__(self, tiles_df, tiles_dir, output_path, design_path):
        self.tiles_df = tiles_df
        self.tiles_dir = tiles_dir
        self.output_path = output_path
        self.design_path = design_path
        self.filtered_tiles = None
        self.tile_counts = None
        self.tile_counts_with_design = None

    def filter_tiles(self):
        self.filtered_tiles = self.tiles_df[
            (self.tiles_df['rampPoints'] == False) &
            (self.tiles_df['underRamp'] == False) &
            (~self.tiles_df['tileType_image'].isin(['ev1.png', 'ev2.png', 'ev3.png']))
        ]

    def count_tile_images(self):
        counts = self.filtered_tiles['tileType_image'].value_counts()
        self.tile_counts = counts.reset_index().rename(columns={'index': 'tileType_image', 'tileType_image': 'amount'})

    def add_design_column(self):
        self.tile_counts_with_design = self.tile_counts.copy()
        self.tile_counts_with_design['design'] = self.tile_counts_with_design['tileType_image'].apply(
            lambda x: os.path.join(self.tiles_dir, x) if os.path.isfile(os.path.join(self.tiles_dir, x)) else None)

    def save_to_excel_with_images(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tiles"

        ws.cell(row=1, column=1, value=f"デザイン: {self.design_path}")
        ws.cell(row=2, column=1, value="構成されるタイルセット")
        ws.cell(row=15, column=1, value="※立体交差下，坂道タイルは，フィールドデザインを確認し作成してください")

        headers = ['tileType_image', 'design', 'amount']
        for i in range(3):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=3, column=col_idx + (i * 3), value=header)

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        column_widths = {
            'A': 110 / 7, 'B': 52 / 7, 'C': 70 / 7,  
            'D': 110 / 7, 'E': 52 / 7, 'F': 70 / 7,  
            'G': 110 / 7, 'H': 52 / 7, 'I': 70 / 7   
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for row_idx in range(3, 14):
            ws.row_dimensions[row_idx].height = 40

        row_idx = 4
        col_offset = 0
        counter = 0

        for _, row in self.tile_counts_with_design.iterrows():
            image_path = row['design']
            col_start = 1 + col_offset
            ws.cell(row=row_idx, column=col_start, value=row['tileType_image']).border = thin_border
            ws.cell(row=row_idx, column=col_start + 2, value=row['amount']).border = thin_border

            if image_path and os.path.isfile(image_path):
                img = Image(image_path)
                img.height = 50
                img.width = 50
                img_anchor = f"{get_column_letter(col_start + 1)}{row_idx}"
                ws.add_image(img, img_anchor)
            ws.cell(row=row_idx, column=col_start + 1).border = thin_border

            counter += 1
            row_idx += 1
            if counter % 10 == 0:
                row_idx = 4
                col_offset += 3

        for row in ws.iter_rows(min_row=3, max_row=13, min_col=1, max_col=9):
            for cell in row:
                cell.border = thin_border

        wb.save(self.output_path)
        print(f"Data saved to {self.output_path}")

    def process_tiles(self):
        self.filter_tiles()
        self.count_tile_images()
        self.add_design_column()
        self.save_to_excel_with_images()


def process_tileset(input_path):
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    xlsx_output = base_name + '_tileList.xlsx'
    tiles_dir = 'tiles'

    # Process tiles directly to Excel
    tile_processor = TileProcessor(input_path, None)
    tile_processor.process(xlsx_output)

    # Count tiles and create Excel
    image_processor = TileImageProcessor(tile_processor.tiles_df_reordered, tiles_dir, xlsx_output, input_path)
    image_processor.process_tiles()


def main():
    root = tk.Tk()
    root.withdraw()
    json_file = filedialog.askopenfilename(title="Select CMS design format json file", filetypes=[("JSON Files", "*.json")])
    if json_file:
        process_tileset(json_file)
    else:
        print("No file selected.")

if __name__ == "__main__":
    main()
